import os
import logging
import time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
# import MySQLdb
from configparser import ConfigParser
from openpyxl import load_workbook


# -------------------------------------------------------------------------------
# ###############################################################################
# -------------------------------------------------------------------------------
# 函数/过程名称：GetNewReport_P
# 函数/过程的目的：获取最新报告文件
# 假设：无
# 影响：无
# 输入：无
# 返回值：文件全路径
# 创建者：
# 创建时间：2018/06/26
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------
FD = "./reports"
def GetNewReport_P(FileDir=FD):
    #打印目录所在所有文件名（列表对象）
    l = os.listdir(FileDir)
    l.sort(key=lambda fn:os.path.getmtime(FileDir + "\\" + fn))
    f = os.path.join(FileDir,l[-1])
    return f


# -------------------------------------------------------------------------------
# ###############################################################################
# -------------------------------------------------------------------------------
# 类名称：SendEmail_P
# 类的目的：发送文本邮件或发送带附件邮件
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：
# 创建时间：2018/06/26
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------
def SendEmail_P(sender, psw, receiver, smtpserver, report_file, port):
    with open(report_file, "rb") as f:
        mail_body = f.read()
    # 定义邮件内容
    msg = MIMEMultipart()
    body = MIMEText(mail_body, _subtype='html', _charset='utf-8')
    msg['Subject'] = u"自动化测试报告"
    msg["from"] = sender
    msg["to"] = psw
    msg.attach(body)
    # 添加附件
    att = MIMEText(open(report_file, "rb").read(), "base64", "utf-8")
    att["Content-Type"] = "application/octet-stream"
    att["Content-Disposition"] = 'attachment; filename= "report.html"'
    msg.attach(att)
    try:
        smtp = smtplib.SMTP_SSL(smtpserver, port)
    except:
        smtp = smtplib.SMTP()
        smtp.connect(smtpserver, port)
    # 用户名,客户端授权密码
    smtp.login(sender, psw)
    smtp.sendmail(sender, receiver, msg.as_string())
    smtp.quit()
    print('邮件发送成功！')


# -------------------------------------------------------------------------------
# ###############################################################################
# -------------------------------------------------------------------------------
# 类名称：InsertLog_P
# 类的目的：写日志
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：
# 创建时间：2018/06/26
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------
#当前文件路径
cur_path = os.path.dirname(os.path.realpath(__file__))
# log_path是存放日志的路径
log_path = os.path.join(os.path.dirname(cur_path), 'logs')
# 如果不存在这个logs文件夹，就自动创建一个
if not os.path.exists(log_path):
    os.mkdir(log_path)

class InsertLog_P():
    def __init__(self):
        # 文件的命名
        self.logname = os.path.join(log_path, '%s.log'%time.strftime('%Y_%m_%d'))
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.DEBUG)
        # 日志输出格式
        self.formatter = logging.Formatter('[%(asctime)s - %(filename)s line: %(lineno)3d] - %(levelname)s: %(message)s')

    def __console(self, level, message):
        # 创建一个FileHandler，用于写到本地
        # fh = logging.FileHandler(self.logname, 'a')  # 追加模式  这个是python2的
        fh = logging.FileHandler(self.logname, 'a', encoding='utf-8')  # 这个是python3的
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(self.formatter)
        self.logger.addHandler(fh)

        # 创建一个StreamHandler,用于输出到控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        ch.setFormatter(self.formatter)
        self.logger.addHandler(ch)

        if level == 'info':
            self.logger.info(message)
        elif level == 'debug':
            self.logger.debug(message)
        elif level == 'warning':
            self.logger.warning(message)
        elif level == 'error':
            self.logger.error(message)
        # 这两行代码是为了避免日志输出重复问题
        self.logger.removeHandler(ch)
        self.logger.removeHandler(fh)
        # 关闭打开的文件
        fh.close()

    def debug(self, message):
        self.__console('debug', message)

    def info(self, message):
        self.__console('info', message)

    def warning(self, message):
        self.__console('warning', message)

    def error(self, message):
        self.__console('error', message)


# -------------------------------------------------------------------------------
# ###############################################################################
# -------------------------------------------------------------------------------
# 函数/过程名称：GetSkipScripts
# 函数/过程的目的：获取不需要执行的模块名字
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：
# 创建时间：2018/06/26
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------

def GetSkipScripts_P(FilePath):
    try:
        m = []
        wb = load_workbook(FilePath)
        ws = wb['ScriptConfig']
        rowcount = ws.max_row
        print(rowcount)
        print('************************************************************')
        for i in range(2,rowcount+1):
            cellvalue = ws.cell(row=i,column=2).value
            if cellvalue=='False':
                modulename = ws.cell(row=i,column=1).value
                m.append(modulename)
        wb.close()
        return m
    except BaseException as msg:
        log = InsertLog_P()
        log.error(msg)


# -------------------------------------------------------------------------------
# ###############################################################################
# -------------------------------------------------------------------------------
# 函数/GetSkipTestCases
# 函数/过程的目的：获取不需要执行的用例
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：
# 创建时间：2018/06/26
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------

def GetSkipTestCases_P(FilePath):
    try:
	    #创建一个空列表，用来接收不执行用例的名字
        t = []
		#读取Excel文件
        wb = load_workbook(FilePath)
		#获取Excel所有表格名字
        sheels = wb.sheetnames
        #print sheels
		#第一层for循环，用于遍历所有表格
        for i in sheels:
		    #获取表格名字
            ws = wb[i]
			#获取表格已使用行数
            rowcount = ws.max_row
			#第二层for循环，用于遍历表格每一行数据(从第二行开始)
            for j in range(2,rowcount+1):
			    #获取单元格信息
                cellvalue = ws.cell(row=j,column=13).value
				#判断获取到单元格数据的值是否为‘False’
                if cellvalue=='False':
                    testcasename = ws.cell(row=j,column=1).value
                    t.append(testcasename)
        wb.close()
        return t
    except BaseException as msg:
        log = InsertLog_P()
        log.error(msg)


# -------------------------------------------------------------------------------
# ###############################################################################
# -------------------------------------------------------------------------------
# 函数名称：GetExcelCellData_P
# 函数的目的：获取excel文件指定表格对应单元格数据
# 假设：无
# 影响：无
# 输入：无
# 返回值：指定表格对应单元格数据
# 创建者：
# 创建时间：2018/06/26
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
def GetExcelCellData_P(filepath, sheetname, row, clo):
    try:
        wb = load_workbook(filepath)
        ws = wb.get_sheet_by_name(sheetname)
        cellvalue = ws.cell(row=row, column=clo).value
        wb.close()
        return cellvalue
    except BaseException as msg:
        log = InsertLog_P()
        log.error(msg)


# -------------------------------------------------------------------------------
# ###############################################################################
# -------------------------------------------------------------------------------
# 函数名称：GetIniFileData_P
# 函数的目的：获取ini文件对应节点条目数据
# 假设：无
# 影响：无
# 输入：无
# 返回值：对应节点条目数据（元组类型）
# 创建者：
# 创建时间：2018/06/26
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
def GetIniFileData_P(filepath, section, option):
    try:
        data = ConfigParser()  #定义data为解析配置文件模块并生效
        data.read(filepath,'utf-8')
        r = eval(str(data[section][option]))  #eval函数就是实现list、dict、tuple与str之间的转化
        return r
    except BaseException as msg:
        log = InsertLog_P()
        log.error(msg)


# -------------------------------------------------------------------------------
# ###############################################################################
# -------------------------------------------------------------------------------
# 函数名称：ReadMySQLData_P
# 函数的目的：查询MySQL数据库单条数据
# 假设：无
# 影响：无
# 输入：无
# 返回值：查询一行数据（元组类型）
# 创建者：
# 创建时间：2018/06/26
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
# def ReadMySQLData_P(host, port, user, password, db, sql):
#     try:
#         conn = MySQLdb.connect(host=host,\
#                                port=port,\
#                                user=user,\
#                                passwd=password,\
#                                db=db )
#         curs = conn.cursor()
#         curs.execute(sql)
#         r = curs.fetchone()
#         curs.close()
#         conn.close()
#         return r
#     except BaseException as msg:
#         log = InsertLog_P()
#         log.error(msg)


# -------------------------------------------------------------------------------
# ###############################################################################
# -------------------------------------------------------------------------------
# 函数名称：DelectMySQLData_P
# 函数的目的：删除MySQL数据库指定表数据
# 假设：无
# 影响：无
# 输入：无
# 返回值：布尔值（True或False）
# 创建者：
# 创建时间：2018/06/26
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
def DelectMySQLData_P(host, port, user, password, db, table ,sql):
    try:
        conn = MySQLdb.connect(host=host,\
                               port=port,\
                               user=user,\
                               passwd=password,\
                               db=db )
        curs = conn.cursor()
        curs.execute('select * from %s' %table)
        before = curs.rowcount

        curs.execute(sql)
        curs.execute('select * from %s' % table)
        after =  curs.rowcount

        curs.close()
        conn.close()

        flat = None

        if before != after:
            flat = True
        else:
            flat = False

        return flat

    except BaseException as msg:
        log = InsertLog_P()
        log.error(msg)



if __name__ == '__main__':




    data = GetIniFileData_P('../data/loginAPI_test_data.ini', 'test_login_success_001', 'input_data')
    expect_result_error_code = GetIniFileData_P('../data/loginAPI_test_data.ini', 'test_login_success_001', 'expect_result_error_code')
    print(data)
    print(expect_result_error_code)

    # #sql = "delete from p2p_mobile_verify_code where mobile=13812341234"
    # sql = "delete from p2p_user where user_name='test016'"
    # data = DelectMySQLData_P('127.0.0.1', 3306, 'root', 'root', 'p2p','p2p_user', sql)
    # print data

    print (GetIniFileData_P('../data/case_data.ini','TestRegisterAPI','test_register_error_002'))

    print (GetExcelCellData_P(u'../testcase/测试用例.xlsx',u'登录接口',2,1))

    log = InsertLog_P()
    log.debug(u"debug日志输出")
    log.info(u"Info日志输出")
    log.warning(u"warning日志输出")
    log.error(u"error日志输出")

    log = InsertLog_P()
    log.info("---测试开始----")
    log.info("操作步骤1,2,3")
    log.warning("----测试结束----")